import json
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
import os
from openpyxl import Workbook
from collections import Counter
import re
import concurrent.futures
from tqdm import tqdm
import logging

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class WantedCrawler:
    def __init__(self, start_id, end_id):
        self.start_id = start_id
        self.end_id = end_id
        self.success_count = 0
        self.error_count = 0
        self.not_found_count = 0
        self.job_data_list = []
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'Connection': 'keep-alive',
        }
        self.session = requests.Session()

    def random_sleep(self, min_seconds=0.5, max_seconds=1):
        """랜덤한 시간 동안 대기"""
        time.sleep(random.uniform(min_seconds, max_seconds))

    def extract_data(self, soup, job_id):
        """JSON 데이터 추출 시도, 실패 시 기존 방식 사용"""
        try:
            script_tag = soup.find('script', {'id': '__NEXT_DATA__'})
            if script_tag:
                json_data = json.loads(script_tag.string)
                job_data = json_data.get('props', {}).get('pageProps', {}).get('job', {})

                if job_data:
                    return {
                        'id': job_id,
                        'company_name': job_data.get('company', {}).get('name', ''),
                        'position': job_data.get('position', ''),
                        'sub_categories': ', '.join(job_data.get('subCategories', [])),
                        'job_description': job_data.get('jd', ''),
                        'url': f"https://www.wanted.co.kr/wd/{job_id}"
                    }
        except Exception as e:
            logging.error(f"JSON 데이터 추출 실패 (ID: {job_id}): {e}")

        # 기존 방식
        try:
            soup_str = str(soup)
            position = soup_str[soup_str.find('"position":') + 12 : soup_str.find('"reward":') - 2]
            sub_categories = soup_str[soup_str.find('"sub_categories":') + 18 : soup_str.find('"position":') - 2]
            job_description = soup_str[soup_str.find('"jd":') + 5 : soup_str.find('"company_name":') - 2]
            company_name = soup_str[soup_str.find('"company_name":') + 16 : soup_str.find('"lang":') - 2]

            return {
                'id': job_id,
                'company_name': company_name,
                'position': position,
                'sub_categories': sub_categories,
                'job_description': job_description,
                'url': f"https://www.wanted.co.kr/wd/{job_id}"
            }
        except Exception as e:
            logging.error(f"기존 방식 데이터 추출 실패 (ID: {job_id}): {e}")
            return None

    def crawl_job(self, job_id):
        """단일 채용공고 크롤링"""
        url = f"https://www.wanted.co.kr/wd/{job_id}"
        try:
            response = self.session.get(url, headers=self.headers, timeout=10)
            response.raise_for_status()  # HTTP 에러 발생 시 예외 발생

            html = response.text
            soup = BeautifulSoup(html, 'lxml')

            job_data = self.extract_data(soup, job_id)
            if job_data:
                self.success_count += 1
                return job_data
            else:
                self.error_count += 1
                logging.warning(f"데이터 추출 실패 (ID: {job_id})")
                return None

        except requests.exceptions.RequestException as e:
            self.error_count += 1
            logging.error(f"요청 실패 (ID: {job_id}): {e}")
            return None

        except Exception as e:
            self.error_count += 1
            logging.error(f"크롤링 실패 (ID: {job_id}): {e}")
            return None

    def crawl_jobs(self, max_workers=10):
        """병렬 크롤링"""
        logging.info("크롤링 시작...")
        job_ids = range(self.start_id, self.end_id + 1)

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(self.crawl_job, job_id) for job_id in job_ids]
            for future in tqdm(concurrent.futures.as_completed(futures), total=len(job_ids), desc="크롤링 진행률"):
                job_data = future.result()
                if job_data:
                    self.job_data_list.append(job_data)
        logging.info("크롤링 완료!")
        logging.info(f"총 성공: {self.success_count}, 실패: {self.error_count}, 없음: {self.not_found_count}")

    def save_to_excel(self):
        """데이터를 Excel 파일로 저장"""
        if not self.job_data_list:
            logging.warning("저장할 데이터가 없습니다.")
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'wanted_jobs_{timestamp}.xlsx'
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

        wb = Workbook()
        ws = wb.active

        # 헤더 설정
        headers = ["ID", "회사이름", "직무", "유사직무", "채용내용", "URL"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 데이터 저장
        for row, job_data in enumerate(self.job_data_list, 2):
            ws.cell(row=row, column=1, value=job_data['id'])
            ws.cell(row=row, column=2, value=job_data['company_name'])
            ws.cell(row=row, column=3, value=job_data['position'])
            ws.cell(row=row, column=4, value=job_data['sub_categories'])
            ws.cell(row=row, column=5, value=job_data['job_description'])
            ws.cell(row=row, column=6, value=job_data['url'])

        wb.save(filepath)
        logging.info(f"데이터 저장 완료: {filepath}")

    def analyze_job_data(self):
        """채용공고 데이터 분석"""
        frontend_count = 0
        backend_count = 0
        frontend_keywords = Counter()
        backend_keywords = Counter()

        # 키워드 정의
        frontend_keywords_list = ['frontend', '프론트엔드', 'front-end', '프론트', 'react', 'vue', 'angular', 'javascript', 'js', 'html', 'css', '웹', 'ui', 'ux']
        backend_keywords_list = ['backend', '백엔드', 'back-end', '백엔드', 'java', 'spring', 'python', 'django', 'flask', 'node', 'express', 'php', 'laravel', '서버', 'api']

        for job in self.job_data_list:
            title = job['position'].lower()
            description = job['job_description'].lower()

            # 프론트엔드/백엔드 분류
            is_frontend = any(keyword in title or keyword in description for keyword in frontend_keywords_list)
            is_backend = any(keyword in title or keyword in description for keyword in backend_keywords_list)

            if is_frontend:
                frontend_count += 1
                # 프론트엔드 키워드 추출
                words = re.findall(r'\b\w+\b', description)
                frontend_keywords.update(words)
            if is_backend:
                backend_count += 1
                # 백엔드 키워드 추출
                words = re.findall(r'\b\w+\b', description)
                backend_keywords.update(words)

        # 결과 저장
        analysis_result = {
            'total_jobs': len(self.job_data_list),
            'frontend_count': frontend_count,
            'backend_count': backend_count,
            'frontend_keywords': dict(frontend_keywords.most_common(20)),
            'backend_keywords': dict(backend_keywords.most_common(20))
        }

        return analysis_result

    def save_analysis_result(self, analysis_result):
        """분석 결과를 JSON 파일로 저장"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'wanted_analysis_{timestamp}.json'
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(analysis_result, f, ensure_ascii=False, indent=2)

        logging.info(f"분석 결과 저장 완료: {filepath}")
        logging.info("\n분석 결과:")
        logging.info(f"총 채용공고 수: {analysis_result['total_jobs']}")
        logging.info(f"프론트엔드 채용공고 수: {analysis_result['frontend_count']}")
        logging.info(f"백엔드 채용공고 수: {analysis_result['backend_count']}")
        logging.info("\n프론트엔드 주요 키워드:")
        for keyword, count in analysis_result['frontend_keywords'].items():
            logging.info(f"- {keyword}: {count}")
        logging.info("\n백엔드 주요 키워드:")
        for keyword, count in analysis_result['backend_keywords'].items():
            logging.info(f"- {keyword}: {count}")

def main():
    print("=" * 50)
    print("원티드 채용공고 크롤링 시작")
    print("=" * 50)
    
    start_id = 1001
    end_id = 47682

    crawler = WantedCrawler(start_id, end_id)
    crawler.crawl_jobs(max_workers=20)  # max_workers 값 조정 가능
    crawler.save_to_excel()

    analysis_result = crawler.analyze_job_data()
    crawler.save_analysis_result(analysis_result)

    print("=" * 50)
    print("크롤링 완료!")
    print("=" * 50)

if __name__ == "__main__":
    main()
